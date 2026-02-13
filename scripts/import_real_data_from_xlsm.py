from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
XLSM_PATH = ROOT / "CALCETTO_2.0.xlsm"
REAL_DATA_DIR = ROOT / "data"
REPORT_PATH = ROOT / "analysis" / "real_data_import_report.json"

SLOT_COLUMNS = list(range(10, 24))  # J..W
GOAL_COLUMNS = list(range(52, 66))  # AZ..BM
ASSIST_COLUMNS = list(range(66, 80))  # BN..CA


@dataclass
class ImportedMatch:
    row_index: int
    date: datetime
    note: str
    players_rows: list[dict[str, object]]
    team_a_players: int
    team_b_players: int
    score_a: int
    score_b: int


def _clean_player_name(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"", "0"}:
        return ""
    return text


def _clean_goal_value(value: object) -> int:
    if value is None or value == "":
        return 0
    return int(value)


def _read_roster(workbook: openpyxl.Workbook) -> list[str]:
    ws = workbook["GIOCATORI"]
    names: list[str] = []
    seen: set[str] = set()
    for row in range(5, ws.max_row + 1):
        name = _clean_player_name(ws.cell(row, 2).value)
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
    return names


def _read_matches(workbook: openpyxl.Workbook) -> tuple[list[ImportedMatch], list[dict[str, object]], set[str]]:
    ws = workbook["data"]
    matches: list[ImportedMatch] = []
    anomalies: list[dict[str, object]] = []
    names_seen: set[str] = set()

    for row in range(4, ws.max_row + 1):
        match_date = ws.cell(row, 2).value
        if match_date in (None, ""):
            continue

        campo = ws.cell(row, 6).value or ""
        mvp = ws.cell(row, 9).value or ""
        note = f"source=CALCETTO_2.0.xlsm; row={row}; campo={campo}; mvp={mvp}"

        players_rows: list[dict[str, object]] = []
        score_by_team = {"A": 0, "B": 0}

        for idx, col in enumerate(SLOT_COLUMNS):
            team = "A" if idx < 7 else "B"
            player = _clean_player_name(ws.cell(row, col).value)
            if not player:
                continue

            goals = _clean_goal_value(ws.cell(row, GOAL_COLUMNS[idx]).value)
            assists = _clean_goal_value(ws.cell(row, ASSIST_COLUMNS[idx]).value)
            names_seen.add(player)
            score_by_team[team] += goals
            players_rows.append({"team": team, "player": player, "goals": goals, "assists": assists})

        team_a_players = sum(1 for p in players_rows if p["team"] == "A")
        team_b_players = sum(1 for p in players_rows if p["team"] == "B")
        score_a_cell = _clean_goal_value(ws.cell(row, 7).value)
        score_b_cell = _clean_goal_value(ws.cell(row, 8).value)

        if team_a_players == 0 or team_b_players == 0:
            anomalies.append(
                {
                    "type": "skipped_match_missing_team",
                    "row": row,
                    "date": str(match_date),
                    "team_a_players": team_a_players,
                    "team_b_players": team_b_players,
                }
            )
            continue

        if score_by_team["A"] != score_a_cell or score_by_team["B"] != score_b_cell:
            anomalies.append(
                {
                    "type": "score_mismatch",
                    "row": row,
                    "date": str(match_date),
                    "score_from_slots": {"A": score_by_team["A"], "B": score_by_team["B"]},
                    "score_from_sheet": {"A": score_a_cell, "B": score_b_cell},
                }
            )

        if team_a_players != 5 or team_b_players != 5:
            anomalies.append(
                {
                    "type": "non_5v5_match",
                    "row": row,
                    "date": str(match_date),
                    "team_a_players": team_a_players,
                    "team_b_players": team_b_players,
                }
            )

        matches.append(
            ImportedMatch(
                row_index=row,
                date=match_date,
                note=note,
                players_rows=players_rows,
                team_a_players=team_a_players,
                team_b_players=team_b_players,
                score_a=score_a_cell,
                score_b=score_b_cell,
            )
        )

    return matches, anomalies, names_seen


def _reset_real_data_dir() -> None:
    REAL_DATA_DIR.mkdir(parents=True, exist_ok=True)
    matches_dir = REAL_DATA_DIR / "matches"
    matches_dir.mkdir(parents=True, exist_ok=True)

    for file in matches_dir.glob("*.xlsx"):
        file.unlink()

    (REAL_DATA_DIR / "deleted_matches.json").write_text("[]", encoding="utf-8")


def _write_players(players: list[str]) -> None:
    df = pd.DataFrame([{"id": idx + 1, "name": name} for idx, name in enumerate(players)])
    df.to_csv(REAL_DATA_DIR / "players.csv", index=False)


def _write_matches(matches: list[ImportedMatch]) -> None:
    matches_dir = REAL_DATA_DIR / "matches"
    for idx, match in enumerate(sorted(matches, key=lambda m: (m.date, m.row_index)), start=1):
        file_name = f"{match.date.strftime('%Y-%m-%d')}__real_{idx:03d}__r{match.row_index}.xlsx"
        file_path = matches_dir / file_name

        meta_df = pd.DataFrame(
            [
                {"key": "date", "value": match.date.strftime("%Y-%m-%d")},
                {"key": "note", "value": match.note},
                {"key": "goals_a", "value": match.score_a},
                {"key": "goals_b", "value": match.score_b},
            ]
        )
        players_df = pd.DataFrame(match.players_rows)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            meta_df.to_excel(writer, sheet_name="meta", index=False)
            players_df.to_excel(writer, sheet_name="players", index=False)


def main() -> None:
    if not XLSM_PATH.exists():
        raise FileNotFoundError(f"File not found: {XLSM_PATH}")

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=True)
    roster_names = _read_roster(wb)
    matches, anomalies, names_in_matches = _read_matches(wb)

    undefined_in_matches = sorted(name for name in names_in_matches if name not in set(roster_names))
    all_players = sorted(set(roster_names).union(names_in_matches), key=lambda n: n.lower())

    _reset_real_data_dir()
    _write_players(all_players)
    _write_matches(matches)

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    report = {
        "source_file": str(XLSM_PATH.name),
        "players_in_roster": len(roster_names),
        "players_in_matches": len(names_in_matches),
        "players_total_written": len(all_players),
        "undefined_players_in_matches": undefined_in_matches,
        "matches_written": len(matches),
        "anomalies": anomalies,
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Imported players: {len(all_players)}")
    print(f"Imported matches: {len(matches)}")
    print(f"Undefined players in matches: {len(undefined_in_matches)} -> {undefined_in_matches}")
    print(f"Anomalies found: {len(anomalies)}")
    print(f"Report written: {REPORT_PATH}")


if __name__ == "__main__":
    main()
